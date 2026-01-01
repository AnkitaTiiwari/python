import openpyxl

wb = openpyxl.load_workbook(filename='example.xlsx')

# print(type(wb))
# print(wb.sheetnames)

# sheetname = wb['Sheet3']
# print(sheetname)

# print(sheetname.title)
# activeSheet = wb.active
# print(activeSheet)

# sheetName = wb['Sheet1']
# print(sheetName['A1'].value)

# c = sheetName['B1']
# print(c.value)
# print(c.row)
# print(c.column)
# print(c.coordinate)

# sheetName.cell(row=3, column=3)
# print(sheetName.cell(row=3, column=3).value)


# sheetName = wb['Sheet1']
# for i in range(1,8): #raw
#     for j in range(1,4): #column
#         print(sheetName.cell(row=i,column=j).value)

#Converting column letters to numbers
# from openpyxl.utils import get_column_letter,column_index_from_string
# print(get_column_letter(1))
# print(column_index_from_string('B'))

# sheet = wb['Sheet1']
# print(get_column_letter(sheet.max_column))
# print(sheet.max_column)
# print(sheet.max_row)
# print(column_index_from_string('AA'))

#Getting rows and columns
sheetname = wb['Sheet1']
# a = tuple(sheetname['A1':'C7'])
# print(a)

# for rowOfCellObjects in sheetname['A1':'C7']:
#     for cellObject in rowOfCellObjects:
#         #print(cellObject.coordinate,cellObject.value)
#         print(cellObject.value)
#     print('------EOL------')

#covert into list
a = list(sheetname.columns)[1]
print(a)

for cellobj in list(sheetname.columns)[1]:
    print(cellobj.value)

