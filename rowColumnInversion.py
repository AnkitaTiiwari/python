import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
for i in range(1, sheet.max_row + 1):
    for j in range(1, i + 1):
        temp = sheet.cell(row=i, column=j).value
        sheet.cell(row=i, column=j).value = sheet.cell(row=j, column=i).value
        sheet.cell(row=j, column=i).value = temp
wb.save('rowColumnInversion_output.xlsx')   