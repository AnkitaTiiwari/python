import openpyxl

wb = openpyxl.Workbook()
#print(wb.sheetnames)

sheet = wb.active
print(sheet.title)

sheet.title = "MySheet"
wb.create_sheet('newSheet')
wb.create_sheet(index=2 , title='Middle Sheet')
print(wb.sheetnames)
del wb['MySheet']

#writing into cells

sheet = wb['Middle Sheet']
sheet['A1'] = "Name"
sheet['B1'] = "Place"

wb.save("myworkbook.xlsx")
