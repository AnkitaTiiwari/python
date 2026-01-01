import openpyxl
#from openpyxl.styles import Font

wb = openpyxl.load_workbook("produceSales.xlsx")
sheet = wb.active

listToUpdate = {
    'Garlic':3.07,
    'Celery':1.19,
    'Lemon':1.27

}

#styledFont = Font(size=24,italic=True,bold=True)

for row in range(2, sheet.max_row + 1):
    produceName = sheet.cell(row=row,column=1).value
    #print(produceName)
    if produceName in listToUpdate:
        print(produceName)
        #sheet.cell(row=row,column=2).font = styledFont
        sheet.cell(row=row,column=2).value = listToUpdate[produceName]
        print(sheet.cell(row=row,column=2).value)
        

wb.save('UpdatedProduceSales.xlsx')        